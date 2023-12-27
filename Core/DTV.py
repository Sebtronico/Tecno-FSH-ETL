from RsInstrument import *
import os
import datetime
import time
import subprocess
import pyautogui
import pytesseract
from pytesseract import Output
import openpyxl
from openpyxl.drawing.image import Image

class DTV(RsInstrument):
    # Inicialización de la clase
    def __init__(self, ip_adress: str = '', id_query: bool = True, reset: bool = False, options: str = None, direct_session: object = None):
        self.ip_adress = ip_adress
        if ip_adress == '':
            resource_name = RsInstrument.list_resources("?*")[0]
            ip_adress = resource_name.split(sep=':')[2]
        else:
            resource_name = f'TCPIP::{ip_adress}::INSTR'
        super().__init__(resource_name, id_query, reset, options, direct_session)
        self.write('SYST:DISP:UPD ON')

    # Nombre que se le dará a la carpeta de soportes
    folder_name = f'.\\Soportes_tv_digital'

    # Contador para nombrar las carpetas, en caso de que ya esté creada
    count = 0

    # Cargue de libro de excel Formato Registro Monitoreo In Situ TDT
    wb = openpyxl.load_workbook(filename = '.\\Formatos\\FOR_Registro Monitoreo In Situ TDT_V5.xlsx')
    consPto  = wb[wb.sheetnames[0]] # Creación de objeto para control de la hoja "Consolidado Punto".
    template = wb[wb.sheetnames[2]] # Creación de objeto para control de la hoja de plantilla "CHXX"

    # Definición del diccionario con las frecuencias centrales de los canales
    tvTable = {
        2:  57,   3: 63,   4: 69,   5: 79,   6: 85,   7: 177,  8: 183,
        9:  189, 10: 195, 11: 201, 12: 207, 13: 213, 14: 473, 15: 479,
        16: 485, 17: 491, 18: 497, 19: 503, 20: 509, 21: 515, 22: 521,
        23: 527, 24: 533, 25: 539, 26: 545, 27: 551, 28: 557, 29: 563,
        30: 569, 31: 575, 32: 581, 33: 587, 34: 593, 35: 599, 36: 605,
        38: 617, 39: 623, 40: 629, 41: 635, 42: 641, 43: 647, 44: 653,
        45: 659, 46: 665, 47: 671, 48: 677, 49: 683, 50: 689, 51: 695
    }
    

    # Creación de la carpeta de soportes de la medida
    def create_folder(self):
        try:
            os.mkdir(self.folder_name)
        except FileExistsError:
            now = datetime.datetime.now()
            self.folder_name = f'{self.folder_name}_{now.hour}_{now.minute}_{now.second}'
            self.create_folder()


    # Generación de captura de pantalla *.png y envío al pc
    def screenshot(self, channel, path, image_name):
        img_path_instr = 'C:\\R_S\\instr\\user\\self.screenshot.png' # Ruta y nombre del self.screenshot en el instrumento
        img_path_pc = f'{path}\\{self.tvTable[channel]}{image_name}.png' # Ruta y nombre del archivo exportado

        self.write_str_with_opc("HCOP:DEV:LANG PNG") # Definición del formato de la imagen
        self.write_str_with_opc(f"MMEM:NAME '{img_path_instr}'") # Creación de la imagen en la memoria
        self.write_str_with_opc("HCOP:IMM") # Captura de la pantalla
        self.query_opc()

        self.read_file_from_instrument_to_pc(img_path_instr, img_path_pc) # Transferencia del archivo al PC
        self.write_str_with_opc(f"MMEMory:DELete '{img_path_instr}'")  # Se elimina el archivo de la memoria del instrumento


    # Función de generación del archivo de datos *.dat
    def get_data_file(self, channel, path):
        file_path_instr = 'C:\\R_S\\instr\\user\\datfile.dat' # Ruta y nombre del archivo en el instrumento
        file_path_pc = f'{path}\\{self.tvTable[channel]}.dat' # Ruta y nombre del archivo exportado

        self.write_str_with_opc(f"MMEM:STOR:TRAC 1,'{file_path_instr}'") # Captura del archivo
        self.query_opc()

        self.read_file_from_instrument_to_pc(file_path_instr, file_path_pc) # Transferencia del archivo al PC
        self.write_str_with_opc(f"MMEMory:DELete '{file_path_instr}'")  # Se elimina el archivo de la memoria del instrumento


    # Función de configuración del modo analizador de espectro (Medida de potencia)
    def setup_spectrum_analyzer(self, input, transducers):
        
        self.write_str_with_opc('INST SAN') # Configura el instrumento al modo "Spectrum Analyzer"
        for transducer in transducers:
            self.write_str_with_opc(f"CORR:TRAN:SEL '{transducer}'") # Selecciona el transductor suministrado por el usuario.
            self.write_str_with_opc('CORR:TRAN ON') # Activa el transductor seleccionado
        self.write_str_with_opc('INP:PRES:STAT ON') # Enciende el preselector.
        self.write_str_with_opc('INP:GAIN:STAT OFF') # Apaga el preamplificador.
        self.write_str_with_opc(f'INP:IMP {input}') # Selecciona la entrada según la entrada de la función.
        self.write_str_with_opc('DISP:TRAC1:MODE AVER') # Selecciona el modo de traza "average" para la traza 1.
        self.write_str_with_opc('DET RMS') # Selecciona el detector "RMS"
        self.write_str_with_opc('CALC:MARK:FUNC:POW:PRES NONE') # Configuración de sin estándar para la medición de potencia en modo ACP 
        self.write_str_with_opc('CALC:MARK:FUNC:POW:SEL ACP') # Activa la medición de potencia absoluta.
        self.write_str_with_opc('POW:ACH:ACP 0') # Configura el número de canales adyacentes a 0.
        self.write_str_with_opc('POW:ACH:BWID 5.830 MHz') # Configuración del ancho de banda a 5.830 MHz.
        self.write_str_with_opc('CALC:MARK:AOFF') # Desactiva todos los marcadores. 
        self.write_str_with_opc('FREQ:SPAN 8 MHz') # Configuración del Span
        self.write_str_with_opc('BAND:RES 30 kHz') # Configuración de resolution bandwidth
        self.write_str_with_opc('BAND:VID 300 kHz') # Configuración de video bandwidth
        self.write_str_with_opc('SWE:TIME 500 ms') # Configuración de  sweeptime
        self.write_str_with_opc('INP:ATT 0 dB') # Configuración de la atenuación


    # Función de medición del modo analizador de espectro (Medida de potencia)
    def measurement_spectrum_analyzer(self, channel):

        self.channel_folder_name = f'{self.folder_name}\\CH_{channel}'

        # Generación de subcarpeta por cada canal
        try:
            os.mkdir(self.channel_folder_name)
        except FileExistsError:
            self.channel_folder_name = f'{self.channel_folder_name}_1'
            os.mkdir(self.channel_folder_name)

        self.write_str_with_opc(f'FREQ:CENT {self.tvTable[channel]} MHz') # Configuración de la frecuencia central.
        self.write_str_with_opc(f'INIT:CONT OFF') # Desactiva la medición contínua
        self.write_str_with_opc(f'SWE:COUN 10') # Configuración del conteo de barridos a 10.
        self.write_str_with_opc(f'INIT;*WAI') # Inicia la medición y aguarda hasta que se complete el número de barridos seleccionado.
        
        # Obtención de la potencia de canal
        channel_power = self.query_bin_or_ascii_float_list('CALC:MARK:FUNC:POW:RES? ACP')[0] # Lectura del nivel de potencia.
        
        if channel_power > 83.75: # Si la potencia es mayor a 99 (Posible señal de saturación u overload)
            waveform = self.query_bin_or_ascii_float_list('TRAC? TRACE1') # Query of the trace in screen
            maximum = max(waveform)
            minimum = min(waveform)
            self.write_str_with_opc('INP:PRES:STAT OFF') # Apaga el preselector.
            self.write_str_with_opc(f'DISP:TRAC:Y:RLEV {maximum + 10}') # Apaga el preselector.
            self.write_str_with_opc(f'DISP:TRAC:Y:RLEV {maximum - minimum + 10}') # Fija el reference level en 83.75 (valor máximo)
        channel_power = round(float(channel_power),2) # Conversión de tipo string a tipo float y redondeo a dos cifras decimales
        self.screenshot(channel, self.channel_folder_name, '') # Toma de la captura de pantalla
        self.write_str_with_opc('FREQ:SPAN 5.830 MHz') # Configuración del Span
        self.get_data_file(channel, self.channel_folder_name)

        self.write_str_with_opc(f'INIT:CONT ON') # Activa la medición contínua

        return channel_power
    
    
    # Función de configuración del modo analizador de tv
    def setup_tv_analyzer(self, channel):
        self.write_str_with_opc('INST CATV')  # Entrar al modo TV / Radio Analyzer / Receiver.
        self.write_str_with_opc('CONF:DTV:MEAS DSP')  # Selecciona la ventana Spectrum
        self.write_str_with_opc('SYST:POS:GPS:DEV PPS2')  # Para que muestre las coordenadas en las imágenes
        self.write_str_with_opc('DISP:MEAS:OVER:GPS:STAT ON')  # Para que muestre las coordenadas en las imágenes
        self.write_str_with_opc('DTV:BAND:CHAN B6MHz') # Configura el ancho de banda del canal de TV a 6 MHz.
        self.write_str_with_opc('DDEM:ISSY TOL') # Configura el modo 'Tolerant' para ISSY processing.
        self.write_str_with_opc('UNIT:POW DBUV') # Configura la unidad de medida por defecto a dBuV.
        self.write_str_with_opc('FREQ:SPAN 10 MHz') # Configura el Span.
        self.write_str_with_opc('DISP:TRAC:Y:RLEV 80') # Configura el expected level.
        self.write_str_with_opc('DISP:TRAC:Y 80 dB') # Configura el range log.
        self.write_str_with_opc('INP:ATT 0 dB') # Configura la atenuación.
        self.write_str_with_opc(f'FREQ:RF {self.tvTable[channel]} MHz') # Configuración de la frecuencia central.

    
    # Medición en el modo 'Spectrum'
    def measurement_spectrum(self, channel, path):

        self.write_str_with_opc('CONF:DTV:MEAS DSP')  # Selecciona la ventana Spectrum
        self.write_str_with_opc('CONF:DTV:MEAS:OOB OFF') # Desactiva todos los modos de medida dentro del modo Spectrum.
        self.write_str_with_opc('CONF:DTV:MEAS:SATT ON') # Activa el modo de función "Shoulders".
        self.write_str_with_opc(f'INIT:CONT OFF') # Desactiva la medición contínua
        self.write_str_with_opc(f'SWE:COUN 10') # Configuración del conteo de barridos a 10.
        self.write_str_with_opc(f'INIT;*WAI') # Inicia la medición y aguarda hasta que se complete el número de barridos seleccionado.

        self.screenshot(channel, path, '_010') # Toma de la captura de pantalla


    # Medición en el modo 'Overview'
    def measurement_overview(self, channel, path):

        # Tabla de conversión de los valores FEC.
        fecTable = {
            'R1_2':'1/2',
            'R3_5':'3/5',
            'R2_3':'2/3',
            'R3_4':'3/4',
            'R4_5':'4/4',
            'R5_6':'5/6',
            '...' :'NA'
        }

        # Tabla de conversión de los valores de intervalo de guardas.
        GINTervaltable = {
            'G1_4'   :'1/4',
            'G19_128':'19/28',
            'G1_8'   :'1/8',
            'G19_256':'19/256',
            'G1_16'  :'1/16',
            'G1_32'  :'1/32',
            'G1_128' :'1/128',
            '---'    :'NA'
        }

        # Medición en la ventana Overview
        self.write_str_with_opc('CONF:DTV:MEAS OVER')
        self.write_str_with_opc('DISP:ZOOM:OVER BERLdpc')  # Hacer zoom a la variable BER bef. LDPC.

        t = time.time()
        
        BERLdpc = self.query_with_opc('CALC:DTV:RES? BERLdpc') # Obtención de la variable BER bef. LDPC en tipo string
        while BERLdpc == '---' and time.time() - t <= 30: # Espera durante 30 segundos a que haya una lectura.
            BERLdpc = self.query_with_opc('CALC:DTV:RES? BERLdpc')
        if BERLdpc == '---':
            BERLdpc = 'No disponible'
        else:
            BERLdpc = float(BERLdpc) # Conversión de tipo str a float.
            
        self.screenshot(channel, path, '_003') # Toma de captura de pantalla

        # Medición en la ventana L1 pre signalling.
        self.write_str_with_opc('CONF:DTV:MEAS L1PRe')
        GINTerval = GINTervaltable[self.query_with_opc('CALC:DTV:RES? GINTerval')] # Obtención de la variable intervalo de guardas
        try:
            PLPCodeRate = fecTable[self.query_with_opc('CALC:DTV:RES:L1Post? DPLP').split(sep=',')[4]] # Obtención de la variable FEC
        except IndexError:
            PLPCodeRate = '---'
        self.screenshot(channel, path, '_004') # Toma de captura de pantalla
        
        # Medición en la ventana L1 post signalling 1.
        self.write_str_with_opc('CONF:DTV:MEAS L1P1')
        self.screenshot(channel, path, '_005') # Toma de captura de pantalla.
        
        # Medición en la ventana L1 post signalling 2.
        self.write_str_with_opc('CONF:DTV:MEAS L1P2')
        self.screenshot(channel, path, '_006') # Toma de captura de pantalla.

        # Medición en la ventana L1 post signalling 3.
        self.write_str_with_opc('CONF:DTV:MEAS L1P3')
        self.screenshot(channel, path, '_011') # Toma de captura de pantalla.
        
        return BERLdpc, GINTerval, PLPCodeRate
    

    # Medición en el modo 'Modulation Analysis'
    def measurement_modulation_analysis(self, channel, path):
        
        # Tabla de conversión de los valores de modulación.
        modTable = {
            'QAM16' :'16QAM',
            'QAM64' :'64QAM',
            'QAM256':'256QAM',
            'QPSK'  :'QPSK',
            'BPSK'  :'BPSK',
            '---'   :'NA'
        }

        # Tabla de conversión de los valores de FFT.
        FFTModeTable = {
            'F2K' :'2k',
            'F4K' :'4k',
            'F8K' :'8k',
            'F1K' :'1k',
            'F8KE':'8k ext',
            'F16K':'16k',
            'F16E':'16k ext',
            'F32K':'32k',
            'F32E':'32k ext',
            '---' :'NA'
        }

        # Medición en la ventana 'Modulation errors'
        self.write_str_with_opc('CONF:DTV:MEAS MERR')  # Selecciona la ventana Modulation errors
        self.write_str_with_opc('DISP:ZOOM:MERR MRPLp') # Zoom a la ariable MER (PLP, RMS)
        
        t = time.time()

        MRPLp = self.query_with_opc('CALC:DTV:RES? MRPLp') # Obtención de la variable MER (PLP, RMS) en tipo str
        while MRPLp == '---' and time.time() - t <= 30: # Espera un 30 segundos hasta obtener una medida válida
            MRPLp = self.query_with_opc('CALC:DTV:RES? MRPLp')
        if MRPLp == '---':
            MRPLp = 'No disponible'
        else:
            MRPLp = round(float(MRPLp),1) # Conversión de str a float y redondeo a 1 cifra decimal.

        self.screenshot(channel, path, '_002') # Toma de captura de pantalla.

        # Medición en la ventana de constelación.
        self.write_str_with_opc('CONF:DTV:MEAS CONS')
        time.sleep(3)
        try:
            cons = modTable[self.query_with_opc('CALC:DTV:RES:L1Post? DPLP').split(sep=',')[1]] # Obtención de la variable modulación.
        except IndexError:
            cons = '---'

        FFTMode = FFTModeTable[self.query_with_opc('CALC:DTV:RES? FFTMode')] # Obtención de la variable FFT.
        self.screenshot(channel, path, '_001') # Toma de captura de pantalla.

        # Medición en la ventana MER vs Carrier.
        self.write_str_with_opc('CONF:DTV:MEAS MERFrequency')
        time.sleep(5)
        self.screenshot(channel, path, '_009') # Toma de captura de pantalla.

        return MRPLp, cons, FFTMode
    

    # Medición en el modo 'Channel Analysis'
    def measurement_channel_analysis(self, channel, path):

        # Medición la ventana Echo Pattern.
        self.write_str_with_opc('CONF:DTV:MEAS EPATtern')
        self.write_str_with_opc('DISP:LIST:STATE OFF') # Desactiva la vista de la lista.

        time.sleep(7) ## Corregir si se puede obtener la traza

        PPATtern = self.query_with_opc('CALC:DTV:RES:L1PR? PPATtern') # Obtención de la variable patrón de pilotos.
        self.screenshot(channel, path, '_007') # Toma de captura de pantalla.
        self.write_str_with_opc('DISP:LIST:STATE ON') # Aciva la vista de la lista.
        self.screenshot(channel, path, '_008') # Toma de captura de pantalla. 
        return PPATtern
    

    # Medición en el modo 'TS Analyzer'
    def measurement_ts_analyzer(self, channel, path):

        # Selecciona la ventana de constelación (necesario para correcta medición del TS Analyzer).
        self.write_str_with_opc('CONF:DTV:MEAS CONS')

        time.sleep(5)
        self.write_str_with_opc('INST TSAN') # Configura el instrumento al modo "TS Analyzer"

        # Apertura del escritorio remoto
        subprocess.run(f"cmdkey /generic:{self.ip_adress} /user:instrument /pass:894129")
        subprocess.Popen(f"mstsc /v:{self.ip_adress} /f", close_fds=True)

        # Se espera hasta encontrar la foto del ETL que semuestra en la ventana Site View
        while True:
            if pyautogui.locateOnScreen('.\\Imagenes\\ETL.png', region = (300, 80, 360, 150), grayscale = True, confidence = 0.99) is not None:
                pyautogui.click(300,80)
                time.sleep(3)
                break
        
        # Se elimina el key para la apertura del escritorio remoto
        subprocess.run(f"cmdkey /delete:{self.ip_adress}")

        # Toma captura de pantalla del diagrama de pastel
        img = pyautogui.screenshot(region=(60,25, 700, 505))
        img.save(f"{path}\TS_{self.tvTable[channel]}_001.png")

        pyautogui.moveTo(62, 220) # Para desplegar el menu de las carpetas
        time.sleep(1)
        pyautogui.moveTo(256, 220) # Al borde derecho de la ventana que se despliega
        pyautogui.click()
        pyautogui.drag(50, 0, 0.5, button='left') # Arrastrar hasta el borde
        time.sleep(1)
        pyautogui.click(170, 215)
        time.sleep(1)

        # Toma campura de pantalla del diagrama de pastel con las carpetas desplegadas
        img = pyautogui.screenshot(region=(60,25, 700, 505))
        img.save(f"{path}\TS_{self.tvTable[channel]}_002.png")

        time.sleep(1)

        pyautogui.click(100, 35) # View
        pyautogui.click(156, 78) # Monitoring
        pyautogui.click(271, 88) # Statis & Log
        time.sleep(1)
        pyautogui.click(678, 70) # Control
        pyautogui.click(559, 150) # Clear statics & log of all inputs

        # Espera de un minuto para tomar la captura Statis & log
        time.sleep(60)

        # Toma campura de pantalla de la ventana Statis & log
        img = pyautogui.screenshot(region=(60,25, 700, 505))
        img.save(f"{path}\TS_{self.tvTable[channel]}_003.png")

        pyautogui.click(195, 60) # Bit rate
        time.sleep(1)

        # Toma captura de pantalla de la ventana Bit rate
        img = pyautogui.screenshot(region=(60,25, 700, 505))
        img.save(f"{path}\TS_{self.tvTable[channel]}_004.png")

        pyautogui.click(277, 60) # Table repetition
        time.sleep(1)

        # Toma captura de pantalla de la ventana Table repetition
        img = pyautogui.screenshot(region=(60,25, 700, 505))
        img.save(f"{path}\TS_{self.tvTable[channel]}_005.png")

        # Devuelve la vista al diagrama de pastel para próxima medida
        pyautogui.click(100, 35) # View
        pyautogui.click(153, 57) # Topology
        pyautogui.click(255, 61) # Site View
    

    # Medición en el modo 'TxCheck'
    def measurement_txcheck(self, plp):
        
        # Función para copiar en portapapeles
        def copy2clip(txt):
            cmd='echo '+txt.strip()+'|clip'
            return subprocess.check_call(cmd, shell=True)
        
        # Llevar el dispositivo a modo local para activar el botón mode
        self.write_str_with_opc('INST CATV') # Configura el instrumento al modo "Analizador de Tv"
        time.sleep(1)
        pyautogui.click(650,460)


        pytesseract.pytesseract.tesseract_cmd = f'C:\\Users\\{os.getlogin()}\\AppData\\Local\\Programs\\Tesseract-OCR\\tesseract'
        while True: # Se espera hasta encontrar la foto del ETL que semuestra en la ventana Site View
            try:
                if pyautogui.locateCenterOnScreen('./Imagenes/Mode.png', grayscale = True, confidence = 0.99) is not None:
                    x, y = pyautogui.locateCenterOnScreen('./Imagenes/Mode.png', grayscale = True, confidence = 0.99)
                    pyautogui.click(x,y)
                    time.sleep(1)
                    print('Imagen encontrada')
                    break
            except:
                pass

        img = pyautogui.screenshot(region=(180, 170, 300, 170))
        palabra = 'TxCheck'
        d = pytesseract.image_to_data(img, lang="spa+eng", output_type=Output.DICT)
        n_boxes = len(d['level'])
        for i in range(n_boxes):
            x, y, text = (d['left'][i], d['top'][i], d['text'][i])
            if text == palabra:
                pyautogui.click(x+180,y+170)

        while True: # Se espera hasta encontrar la foto del ETL que semuestra en la ventana Site View
            try:
                if pyautogui.locateCenterOnScreen('./Imagenes/Logo.png', grayscale = True, confidence = 0.95) is not None:
                    time.sleep(1)
                    print('Logo encontrado')
                    break
            except:
                pass

        pyautogui.click(90,85)
        time.sleep(0.5)
        pyautogui.click(660,220)
        time.sleep(2)
        pyautogui.click(308,55)
        time.sleep(1)

        while not pyautogui.pixelMatchesColor(308, 50, (0, 255, 0)):
            continue

        pyautogui.click(400,120)
        pyautogui.hotkey(['ctrl', 's'])
        time.sleep(1)

        txcName = fr'C:\R_S\instr\user\TxCheck1_{plp}.ETLtxc'

        copy2clip(txcName)

        pyautogui.hotkey(['ctrl', 'v'])
        pyautogui.press('enter')
        time.sleep(5)

        pyautogui.doubleClick(981,596)
        time.sleep(1)
        pyautogui.click(1908,5)
        
        time.sleep(2)

        # self.read_file_from_instrument_to_pc(txcName, path) # Transferencia del archivo al PC
        # self.write_str_with_opc(f"MMEMory:DELete '{txcName}'")  # Se elimina el archivo de la memoria del ETL

        # Cerrar la ventana de escritorio remoto
        pyautogui.moveTo(990,0) # Puntero arriba al centro de la pantalla
        time.sleep(2)
        pyautogui.click(1260,15) # Clic en cerrar
        time.sleep(0.250)
        pyautogui.press('enter') # Aceptar
        time.sleep(0.250)


    # Función que ejecuta la medición según la banda
    def measurement_dtv(self, channel, number_of_plp):

        # Declaración del diccionario de resultados
        dtv_results = dict()

        # Medición para cada PLP
        for i in range(number_of_plp): # Se realiza una rutina de medida por cada PLP encontrado en el canal.
            if number_of_plp == 2:
                plp = i
                self.write_str_with_opc('SENS:DDEM:DECP:MODE MAN') # Configuración de selección manual de PLP.
                self.write_str_with_opc(f'DDEM:DECP:MAN {plp}') # Selecciona el PLP.
            elif number_of_plp == 1:
                self.write_str_with_opc('SENS:DDEM:DECP:MODE AUTO')
                plp = self.query_with_opc('CALC:DTV:RES:L1Post? DPLP').split(sep=',')[0]
                t = time.time()
                while plp == '---' and time.time() - t <= 15:
                    plp = self.query_with_opc('CALC:DTV:RES:L1Post? DPLP').split(sep=',')[0]
            else:
                plp = int(self.query_with_opc('CALC:DTV:RES:L1Post? APLP').split(sep=',')[21*i + 2]) # Lectura del número de PLP.
                self.write_str_with_opc('SENS:DDEM:DECP:MODE MAN') # Configuración de selección manual de PLP.
                self.write_str_with_opc(f'DDEM:DECP:MAN {plp}') # Selecciona el PLP.
            

            self.plp_folder_name = f'{self.folder_name}\\CH_{channel}\\PLP_{plp}'

            # Generación de subcarpeta por cada PLP
            os.mkdir(self.plp_folder_name)
                
            self.measurement_spectrum(channel, self.plp_folder_name) # Medida en la ventana Spectrum.
            BERLdpc, GINTerval, PLPCodeRate = self.measurement_overview(channel, self.plp_folder_name) # Medida y obtención de variables en la ventana Overview.
            MRPLp, cons, FFTMode = self.measurement_modulation_analysis(channel, self.plp_folder_name) # Medida y obtención de variables en la ventana Modulation Analysis.
            PPATtern = self.measurement_channel_analysis(channel, self.plp_folder_name) # Medida y obtención de variable en la ventana Channel Analysis.
            dtv_results[plp] = [MRPLp, BERLdpc, cons, PLPCodeRate, FFTMode, GINTerval, PPATtern] # Obtención de todas las variables necesarias.
            if MRPLp != 'No disponible': # Si el BER logró ser medido
                self.measurement_ts_analyzer(channel, self.plp_folder_name) # Se ejecuta la medición de TS Analyzer, de modo contrario, no.
                self.measurement_txcheck(plp)

        return dtv_results # Diccionario con las variables MER, BER, Modulación, FEC, FFT, Intervalo de guardas y Patrón de pilotos, para cada PLP.
    

    # Función de medición completa
    def measurement(self, input, transducers, channel, number_of_plp):
        self.setup_spectrum_analyzer(input, transducers)
        channel_power = self.measurement_spectrum_analyzer(channel)
        self.setup_tv_analyzer(channel)
        dtv_results = self.measurement_dtv(channel, number_of_plp)

        for key in dtv_results:
            dtv_results[key].insert(0,channel_power) # Agrega la variable potencia de canal al diccionario anterior.

        return dtv_results
    

    # Función para crear y rellenar hojas de los canales
    def fill_excel_channel_sheets(self, channel, dtv_results):

        self.consPto [f'A{6 + self.count}'] = f'CH{channel}'

        self.wb.copy_worksheet(self.template) # Creación de una nueva hoja en el formato, copia de la hoja "Template".
        ws = self.wb[self.wb.sheetnames[-1]] # Ubicación de la nueva hoja al final del formato.
        ws.title = f'CH{channel}' # Modificación del nombre de la nueva hoja a "CH<Numero del canal>"
        ws['Q8'] = self.tvTable[channel] # Diligencia el valor de frecuencia en MHz en la casilla Q8 de la hoja.

        first_plp = list(dtv_results.keys())[0]
        imgs = [ # Carga de capturas de pantalla.
            Image(f'{self.channel_folder_name}\\{self.tvTable[channel]}.png'),
            Image(f'{self.channel_folder_name}\\PLP_{first_plp}\\{self.tvTable[channel]}_002.png'),
            Image(f'{self.channel_folder_name}\\PLP_{first_plp}\\{self.tvTable[channel]}_003.png'),
            Image(f'{self.channel_folder_name}\\PLP_{first_plp}\\{self.tvTable[channel]}_001.png'),
            Image(f'{self.channel_folder_name}\\PLP_{first_plp}\\{self.tvTable[channel]}_007.png'),
            Image(f'{self.channel_folder_name}\\PLP_{first_plp}\\{self.tvTable[channel]}_010.png')
        ]

        for img in imgs: # Modificación del tamaño de las imágenes.
            img.width  = img.width/3.1
            img.height = img.height/3.1

        ws.add_image(imgs[0], 'A44') # Se añade la imagen "ChanelPower" a la casilla A44.
        ws.add_image(imgs[1], 'M44') # Se añade la imagen "ModulationErrors" a la casilla M44.
        ws.add_image(imgs[2], 'Z44') # Se añade la imagen "DigitalOverview" a la casilla Z44.
        ws.add_image(imgs[3], 'A64') # Se añade la imagen "Constelation" a la casilla A62.
        ws.add_image(imgs[4], 'M64') # Se añade la imagen "EchoPattern" a la casilla M62.
        ws.add_image(imgs[5], 'Z64') # Se añade la imagen "Shoulders" a la casilla Z62.

        index = 0
        for key in dtv_results: # Diligenciamiento de parámetros de calidad del servicio en la hoja.
            ws[f'J{19 + index}']  = key
            ws[f'L{19 + index}']  = dtv_results[key][0]
            ws[f'P{19 + index}']  = dtv_results[key][1]
            ws[f'R{19 + index}']  = dtv_results[key][2]
            ws[f'V{19 + index}']  = dtv_results[key][3]
            ws[f'Y{19 + index}']  = dtv_results[key][4]
            ws[f'AA{19 + index}'] = dtv_results[key][5]
            ws[f'AD{19 + index}'] = dtv_results[key][6]
            ws[f'AG{19 + index}'] = dtv_results[key][7]
            index += 1

        self.count += 1

        # Guardado del documento por cada canal medido
        self.wb.save(f'{self.folder_name}\\FormatoDiligenciado.xlsx')


# Ejemplo de uso
if __name__ == '__main__':
    pass